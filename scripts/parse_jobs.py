import openpyxl
import json
import sys

def parse_europe(path):
    jobs = []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Todas as Vagas']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 4: continue  # skip header rows
        if not r or not r[1]: continue
        empresa = str(r[1] or '').strip()
        cargo = str(r[2] or '').strip()
        local = str(r[3] or '').strip()
        salario_mensal = r[4]
        salario_anual = r[5]
        funcao = str(r[6] or '').strip()
        link = str(r[7] or '').strip()
        tipo = str(r[8] or '').strip()
        pais = str(r[9] or '').strip()
        data = str(r[10] or '').strip()
        if not cargo or cargo == 'Cargo': continue
        jobs.append({
            'empresa': empresa, 'cargo': cargo, 'local': local,
            'salario_mensal': str(salario_mensal) if salario_mensal else '',
            'salario_anual': str(salario_anual) if salario_anual else '',
            'funcao': funcao, 'link': link, 'tipo': tipo,
            'pais': pais, 'data': data, 'regiao': 'Europa'
        })
    wb.close()
    return jobs

def parse_asia(path):
    jobs = []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Todas as Vagas']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 4: continue
        if not r or not r[1]: continue
        empresa = str(r[1] or '').strip()
        cargo = str(r[2] or '').strip()
        local = str(r[3] or '').strip()
        salario_mensal = r[4]
        salario_anual = r[5]
        funcao = str(r[6] or '').strip()
        link = str(r[7] or '').strip()
        tipo = str(r[8] or '').strip()
        pais = str(r[9] or '').strip()
        data = str(r[10] or '').strip()
        if not cargo or cargo == 'Cargo': continue
        jobs.append({
            'empresa': empresa, 'cargo': cargo, 'local': local,
            'salario_mensal': str(salario_mensal) if salario_mensal else '',
            'salario_anual': str(salario_anual) if salario_anual else '',
            'funcao': funcao, 'link': link, 'tipo': tipo,
            'pais': pais, 'data': data, 'regiao': 'Asia'
        })
    wb.close()
    return jobs

def parse_usa(path):
    jobs = []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Todas as Vagas']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 3: continue
        if not r or not r[1]: continue
        empresa = str(r[1] or '').strip()
        cargo = str(r[2] or '').strip()
        salario_mensal = r[3]
        salario_anual = r[4]
        funcao = str(r[5] or '').strip() if len(r) > 5 else ''
        local = str(r[6] or '').strip() if len(r) > 6 else ''
        link = str(r[7] or '').strip() if len(r) > 7 else ''
        tipo = str(r[8] or '').strip() if len(r) > 8 else ''
        if not cargo or cargo == 'Cargo': continue
        jobs.append({
            'empresa': empresa, 'cargo': cargo, 'local': local,
            'salario_mensal': str(salario_mensal) if salario_mensal else '',
            'salario_anual': str(salario_anual) if salario_anual else '',
            'funcao': funcao, 'link': link, 'tipo': tipo,
            'pais': 'United States', 'data': '', 'regiao': 'EUA'
        })
    wb.close()
    return jobs

print('Parsing Europe...', flush=True)
eu_jobs = parse_europe('upload/Vagas_Tecnologia_Europa_14987.xlsx')
print(f'  -> {len(eu_jobs)} jobs', flush=True)

print('Parsing Asia...', flush=True)
asia_jobs = parse_asia('upload/Vagas_Tecnologia_Asia_10462.xlsx')
print(f'  -> {len(asia_jobs)} jobs', flush=True)

print('Parsing USA...', flush=True)
usa_jobs = parse_usa('upload/Vagas_Tecnologia_EUA_19591.xlsx')
print(f'  -> {len(usa_jobs)} jobs', flush=True)

all_jobs = eu_jobs + asia_jobs + usa_jobs
print(f'Total: {len(all_jobs)} jobs', flush=True)

# Extract unique functions
functions = sorted(set(j['funcao'] for j in all_jobs if j['funcao']))
print(f'Unique functions: {len(functions)}', flush=True)
for f in functions:
    print(f'  - {f}')

# Save all jobs as JSON
with open('data/all_jobs.json', 'w') as fp:
    json.dump(all_jobs, fp, ensure_ascii=False)
print(f'\nSaved to data/all_jobs.json ({len(all_jobs)} jobs)')
